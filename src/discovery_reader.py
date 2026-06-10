"""
discovery_reader.py - DR Discovery Tool Output Parser
=====================================================
Parses JSON and Excel output from azure-terraform-dr-discovery-tool
and returns a list of DiscoveredResource objects filtered to those
that need module or deployment generation.

SAFETY: Read-only. Never modifies input files.
"""

import json
import logging
from pathlib import Path
from typing import List, Optional, Dict, Any

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from models import (
    DiscoveredResource, PrivateEndpointData, DiagnosticData,
    IdentityData, ServiceType, ResourceNeed
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Status -> ResourceNeed mapping
# ---------------------------------------------------------------------------

# Map DR tool comparison_status values to ResourceNeed enum
STATUS_NEED_MAP: Dict[str, ResourceNeed] = {
    "azure_only": ResourceNeed.AZURE_ONLY,
    "module_available_but_not_instantiated": ResourceNeed.MODULE_AVAILABLE,
    "possible_match": ResourceNeed.POSSIBLE_MATCH,
    "terraform_managed": ResourceNeed.TERRAFORM_MANAGED,
    "tf_only": ResourceNeed.UNKNOWN,
    "unknown": ResourceNeed.UNKNOWN,
    # Legacy / alternate names from older discovery tool versions
    "module_missing": ResourceNeed.MODULE_MISSING,
    "deployment_missing": ResourceNeed.DEPLOYMENT_MISSING,
}

# Resource types that need builder attention
ACTIONABLE_NEEDS = {
    ResourceNeed.AZURE_ONLY,
    ResourceNeed.MODULE_AVAILABLE,
    ResourceNeed.MODULE_MISSING,
    ResourceNeed.DEPLOYMENT_MISSING,
    ResourceNeed.POSSIBLE_MATCH,
}

# Map Azure resource types to ServiceType
RESOURCE_TYPE_SERVICE_MAP: Dict[str, ServiceType] = {
    "microsoft.search/searchservices": ServiceType.AI_SEARCH,
    "microsoft.cognitiveservices/accounts": ServiceType.OPENAI,
    "microsoft.network/privateendpoints": ServiceType.PRIVATE_ENDPOINT,
    "microsoft.insights/diagnosticsettings": ServiceType.DIAGNOSTIC_SETTING,
    "microsoft.authorization/roleassignments": ServiceType.RBAC,
}


# ---------------------------------------------------------------------------
# JSON reader
# ---------------------------------------------------------------------------

def read_json_report(file_path: str) -> List[DiscoveredResource]:
    """
    Parse a JSON report from azure-terraform-dr-discovery-tool.

    Args:
        file_path: Path to the .json report file.

    Returns:
        List of DiscoveredResource objects.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Discovery JSON not found: {file_path}")

    logger.info("Reading JSON report: %s", file_path)
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON in {file_path}: {exc}") from exc

    resources = []

    # Handle both top-level list and nested {"azure_resources": [...]} structure
    raw_resources = []
    if isinstance(data, list):
        raw_resources = data
    elif isinstance(data, dict):
        # Try nested structures from DR tool
        if "azure_resources" in data:
            raw_resources = data["azure_resources"]
        elif "comparison_results" in data:
            # Extract from comparison results
            for cr in data["comparison_results"]:
                if cr.get("azure_resource"):
                    raw = cr["azure_resource"]
                    raw["_comparison_status"] = cr.get("status", "")
                    raw["_risk_level"] = cr.get("risk_level", "")
                    raw["_tf_matches"] = cr.get("tf_matches", 0)
                    raw["_recommended_action"] = cr.get("recommended_action", "")
                    raw["_confidence"] = cr.get("confidence", "")
                    raw_resources.append(raw)
        else:
            raw_resources = [data]

    for raw in raw_resources:
        if not raw:
            continue
        res = _parse_raw_resource(raw)
        resources.append(res)

    logger.info("Parsed %d resources from JSON", len(resources))
    return resources


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def read_excel_report(file_path: str, sheet_name: str = "Azure Resources") -> List[DiscoveredResource]:
    """
    Parse an Excel report from azure-terraform-dr-discovery-tool.

    Reads the 'Azure Resources' sheet (or the sheet specified).
    If a 'Terraform Matches' or 'Risks' sheet is present, it enriches
    the resources with comparison status.

    Args:
        file_path: Path to the .xlsx report file.
        sheet_name: Name of the sheet to read resources from.

    Returns:
        List of DiscoveredResource objects.
    """
    if not EXCEL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel input. Install with: pip install openpyxl"
        )

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Discovery Excel not found: {file_path}")

    logger.info("Reading Excel report: %s (sheet: %s)", file_path, sheet_name)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Read main resources sheet
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {file_path}. Available: {available}"
        )

    ws = wb[sheet_name]
    headers = []
    resources_by_name: Dict[str, DiscoveredResource] = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
                      for i, h in enumerate(row)]
            continue
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))
        res = _parse_excel_row(row_dict)
        if res.name:
            resources_by_name[res.name] = res

    # Enrich from Risks sheet if available
    if "Risks" in wb.sheetnames:
        ws_risks = wb["Risks"]
        risk_headers = []
        for row_idx, row in enumerate(ws_risks.iter_rows(values_only=True)):
            if row_idx == 0:
                risk_headers = [
                    str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
                    for i, h in enumerate(row)
                ]
                continue
            if not any(row):
                continue
            rd = dict(zip(risk_headers, row))
            name = str(rd.get("resource_name", "") or "").strip()
            if name and name in resources_by_name:
                resources_by_name[name].comparison_status = str(
                    rd.get("status", "") or ""
                ).strip()
                resources_by_name[name].risk_level = str(
                    rd.get("risk_level", "") or ""
                ).strip()
                resources_by_name[name].recommended_action = str(
                    rd.get("recommended_action", "") or ""
                ).strip()
                resources_by_name[name].confidence = str(
                    rd.get("confidence", "") or ""
                ).strip()

    resources = list(resources_by_name.values())
    logger.info("Parsed %d resources from Excel", len(resources))
    return resources


# ---------------------------------------------------------------------------
# Private parsers
# ---------------------------------------------------------------------------

def _parse_raw_resource(raw: Dict[str, Any]) -> DiscoveredResource:
    """Parse a raw dict (from JSON) into a DiscoveredResource."""
    res = DiscoveredResource()
    res.name = str(raw.get("name", "") or "").strip()
    res.resource_type = str(raw.get("resource_type", "") or "").strip()
    res.resource_group = str(raw.get("resource_group", "") or "").strip()
    res.subscription_id = str(raw.get("subscription_id", "") or "").strip()
    res.location = str(raw.get("location", "") or "").strip()
    res.sku_name = str(raw.get("sku_name", "") or "").strip()
    res.sku_tier = str(raw.get("sku_tier", "") or "").strip()
    res.tags = dict(raw.get("tags") or {})
    res.public_network_access = str(raw.get("public_network_access", "") or "").strip()

    # Identity
    id_raw = raw.get("identity") or {}
    if id_raw:
        res.identity = IdentityData(
            type=str(id_raw.get("type", "None") or "None"),
            principal_id=str(id_raw.get("principal_id", "") or ""),
            user_assigned_identities=list(id_raw.get("user_assigned_identities") or []),
        )

    # Private endpoints
    for pe_raw in (raw.get("private_endpoints") or []):
        if not pe_raw:
            continue
        dns_list = pe_raw.get("dns_configs") or []
        first_fqdn = dns_list[0].get("fqdn", "") if dns_list else ""
        res.private_endpoints.append(PrivateEndpointData(
            name=str(pe_raw.get("name", "") or ""),
            resource_group=str(pe_raw.get("resource_group", "") or ""),
            vnet_name=str(pe_raw.get("vnet_name", "") or ""),
            subnet_name=str(pe_raw.get("subnet_name", "") or ""),
            subnet_id=str(pe_raw.get("subnet_id", "") or ""),
            private_ip_address=str(pe_raw.get("private_ip_address", "") or ""),
            nic_name=str(pe_raw.get("nic_name", "") or ""),
            connection_state=str(pe_raw.get("connection_state", "") or ""),
            group_ids=list(pe_raw.get("group_ids") or []),
            dns_fqdn=first_fqdn,
            private_link_service_id=str(pe_raw.get("private_link_service_id", "") or ""),
        ))

    # Diagnostic settings
    for ds_raw in (raw.get("diagnostic_settings") or []):
        if not ds_raw:
            continue
        res.diagnostic_settings.append(DiagnosticData(
            name=str(ds_raw.get("name", "") or ""),
            log_analytics_workspace_id=str(ds_raw.get("log_analytics_workspace_id", "") or ""),
            log_analytics_workspace_name=str(ds_raw.get("log_analytics_workspace_name", "") or ""),
            storage_account_id=str(ds_raw.get("storage_account_id", "") or ""),
            log_categories=list(ds_raw.get("log_categories") or []),
            metric_categories=list(ds_raw.get("metric_categories") or []),
        ))

    # Comparison fields (may come from wrapped comparison_results)
    res.comparison_status = str(raw.get("_comparison_status", "") or raw.get("comparison_status", "") or "").strip()
    res.risk_level = str(raw.get("_risk_level", "") or raw.get("risk_level", "") or "").strip()
    res.tf_matches = int(raw.get("_tf_matches", 0) or raw.get("tf_matches", 0) or 0)
    res.recommended_action = str(raw.get("_recommended_action", "") or raw.get("recommended_action", "") or "").strip()
    res.confidence = str(raw.get("_confidence", "") or raw.get("confidence", "") or "").strip()

    # Derive service type and need
    res.service_type = _classify_service_type(res)
    res.need = _classify_need(res)

    return res


def _parse_excel_row(row: Dict[str, Any]) -> DiscoveredResource:
    """Parse a single Excel row dict into a DiscoveredResource."""
    res = DiscoveredResource()
    res.name = str(row.get("name", "") or "").strip()
    res.resource_type = str(row.get("resource_type", "") or "").strip()
    res.resource_group = str(row.get("resource_group", "") or "").strip()
    res.subscription_id = str(row.get("subscription_id", "") or "").strip()
    res.location = str(row.get("location", "") or "").strip()
    res.sku_name = str(row.get("sku", "") or row.get("sku_name", "") or "").strip()
    # Tags: "key=val; key2=val2"
    tags_str = str(row.get("tags", "") or "").strip()
    if tags_str:
        for pair in tags_str.split(";"):
            if "=" in pair:
                k, _, v = pair.strip().partition("=")
                res.tags[k.strip()] = v.strip()
    identity_type = str(row.get("identity_type", "") or "").strip()
    if identity_type:
        res.identity = IdentityData(type=identity_type)
    res.public_network_access = str(row.get("public_network_access", "") or "").strip()
    res.comparison_status = str(row.get("comparison_status", "") or "").strip()
    res.service_type = _classify_service_type(res)
    res.need = _classify_need(res)
    return res


def _classify_service_type(res: DiscoveredResource) -> ServiceType:
    """Derive ServiceType from resource_type and extras."""
    rtype = res.resource_type.lower().strip()
    svc = RESOURCE_TYPE_SERVICE_MAP.get(rtype)
    if svc:
        # Distinguish OpenAI vs AI Foundry by kind
        if svc == ServiceType.OPENAI:
            kind = str(res.extras.get("kind", "") or "").lower()
            if kind.startswith("ai") or kind == "aiservices":
                return ServiceType.AI_FOUNDRY
        return svc
    # Fallback: name-based hints
    name_lower = res.name.lower()
    if "search" in name_lower:
        return ServiceType.AI_SEARCH
    if "openai" in name_lower or "oai" in name_lower:
        return ServiceType.OPENAI
    if "foundry" in name_lower or "aifoundry" in name_lower:
        return ServiceType.AI_FOUNDRY
    return ServiceType.GENERIC


def _classify_need(res: DiscoveredResource) -> ResourceNeed:
    """Determine what type of generation is needed."""
    status = res.comparison_status.lower().strip()
    return STATUS_NEED_MAP.get(status, ResourceNeed.UNKNOWN)


# ---------------------------------------------------------------------------
# Filter helpers
# ---------------------------------------------------------------------------

def filter_actionable(resources: List[DiscoveredResource]) -> List[DiscoveredResource]:
    """
    Filter resources that need module builder attention.
    Excludes already terraform_managed resources.
    """
    actionable = [
        r for r in resources
        if r.need in ACTIONABLE_NEEDS
    ]
    logger.info(
        "%d of %d resources need builder action",
        len(actionable), len(resources)
    )
    return actionable


def read_discovery_file(file_path: str) -> List[DiscoveredResource]:
    """
    Auto-detect format (.json or .xlsx) and parse discovery output.

    Args:
        file_path: Path to DR tool output file.

    Returns:
        All discovered resources from the file.
    """
    path = Path(file_path)
    ext = path.suffix.lower()
    if ext == ".json":
        return read_json_report(file_path)
    elif ext in (".xlsx", ".xls"):
        return read_excel_report(file_path)
    else:
        # Try JSON first, then Excel
        try:
            return read_json_report(file_path)
        except Exception:
            return read_excel_report(file_path)
